#!/usr/bin/env python
# -*- coding:utf-8 -*-
#@Time    :    2019/10/22 0022 22:32
#@Author  :    tb_youth
#@FileName:    FisherScore.py
#@SoftWare:    PyCharm
#@Blog    :    https://blog.csdn.net/tb_youth

from sklearn.datasets import make_multilabel_classification
import numpy as np

from openpyxl import workbook,load_workbook

def read_excel_xlsx(path, sheet_name='Sheet'):
    wb = load_workbook(path)
    sheet = wb[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end='')
        print('')


#随机生成数据
def get_data():
    x,y =  make_multilabel_classification(n_samples=20,n_features=2,\
                                          n_labels=1,n_classes=1,random_state=2)
    # #创建表格
    # wb = workbook.Workbook()
    # #表示动作句柄
    # wa = wb.active
    # for i in range(len(x)):
    #     # print(list(x[i])+list(y[i]))
    #     wa.append(list(x[i])+list(y[i]))
    # wb.save('data.xlsx')
    # # read_excel_xlsx('data.xlsx')


    # x：特征值，y：类别
    # 根据类别分个类
    # 类别1的下标
    index1 = np.array([index for (index,value) in enumerate(y) if value == 0]) # print(index1)
    #类别2的下标
    index2 = np.array([index for (index,value) in enumerate(y) if value == 1])
    c1  = x[index1]
    c2 = x[index2]
    return x,np.array([c1,c2])

def fisher_score():
    #x：整体特征数据，class_p:分类后特征数据
    x,class_p = get_data()
    # #axis=0:纵轴，axis=1：横轴
    # print(np.mean(x,axis=0))
    #特征均值（不分类，每个特征的特征均值）
    total_mean = np.mean(x,axis=0)
    # print(total_mean)
    #根据分类，每个特征的均值
    lst = []
    for item in class_p:
        lst.append(np.mean(item,axis=0))
    # print(lst)
    #分类后，每个特征样本数目
    ce_n = []
    for cl in class_p:
        ce_n.append(len(cl))
    # print(ce_n)
    #每一类：该类样本数 *（该类样本均值-特征均值）^2
    m = []
    for c in lst:
        #print(np.power(np.subtract(c,total_mean),2))
        m.append(np.multiply(ce_n,np.power(np.subtract(c,total_mean),2)))
    #print(m)
    #up_sum：公式上方所表述的和
    up_sum = np.sum(m,axis=0)
    # print(up_sum)
    #每一类每个特征的标准差
    # print(class_p)
    s = []
    for i in range(len(class_p)):
        # print(ce_n[i],np.std(class_p[i],axis=0))
        s.append((ce_n[i])*np.std(class_p[i],axis=0))
    # print(s)
    down_sum = np.sum(s,axis=0)
    # print(down_sum)
    score_lst = np.divide(up_sum,down_sum)
    print('+---------------fisherscore-----------------+')
    print(score_lst)


















if __name__ == '__main__':
    fisher_score()
